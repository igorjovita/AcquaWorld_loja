import os

from openpyxl import load_workbook
import calendar
import streamlit as st
import json
import re


class Excel:
    def __init__(self, caminho_arquivo, wb_modelo, repository_audit_log, wb_modelo_caixa, caminho_caixa):

        self.repository_audit_log = repository_audit_log
        self.caminho = caminho_arquivo
        self.wb = load_workbook(caminho_arquivo)
        self.ws = None
        self.wb_modelo = load_workbook(wb_modelo)
        self.wb_modelo_caixa = load_workbook(wb_modelo_caixa)

        self.wb_caixa = load_workbook(caminho_caixa)

    def controlador(self, selects_pendentes):

        dicionario_formatado = {}
        dicionario_caixa = {}

        for i, select in enumerate(selects_pendentes):
            id = select[0]
            tipo = select[1]
            tabela = select[2]
            dicionario = json.loads(select[3])
            tabelas_juntas = ['reserva', 'pagamentos', 'cliente']
            nome = ''
            if tipo == 'INSERT':
                if tabela in tabelas_juntas:
                    pasta_excel = 'RESERVAS'
                    if tabela == 'reserva' or tabela == 'pagamentos':
                        nome = dicionario['nome_cliente']

                    elif tabela == 'cliente':
                        nome = dicionario['nome']

                    if nome:  # Verifica se nome não é None ou vazio
                        if nome not in dicionario_formatado:
                            dicionario_formatado[nome] = {}

                        for key, value in dicionario.items():
                            if key != 'nome':
                                dicionario_formatado[nome][key] = value
                                dicionario_formatado[nome][f'id{i}'] = id

                if tabela == 'caixa':

                    descricao = dicionario.get('descricao')

                    if descricao:

                        if descricao not in dicionario_caixa:
                            dicionario_caixa[descricao] = {}

                        for key, value in dicionario.items():
                            dicionario_caixa[descricao][key] = value
                            dicionario_caixa[descricao][f'id{i}'] = id



        st.write(dicionario_caixa)
        return dicionario_formatado, dicionario_caixa

    def insert_excel_caixa(self, dicionario):

        for nome, item in dicionario.items():
            descricao = nome

            lista_ids_audit_log = []
            st.write('Abaixo')
            st.write(dicionario)
            st.write('Item')
            st.write(item)

            data = item['data']
            self.selecionar_sheet_por_data(data, 'caixa')

            tipo_movimento = item['tipo_movimento']
            proxima_linha_vazia = None
            coluna_tipo = 0
            coluna_descricao = 0
            coluna_forma_pg = 0
            coluna_valor = 0

            if tipo_movimento == 'ENTRADA':
                coluna_tipo = 2
                coluna_descricao = 3
                coluna_forma_pg = 4
                coluna_valor = 5

                for row in range(9, self.ws.max_row + 2):
                    linha_vazia = True
                    for col in range(2, 5):  # Colunas B (2) a G (7)
                        cell = self.ws.cell(row=row, column=col)
                        if cell.value is not None:
                            linha_vazia = False
                            break
                    if linha_vazia:
                        proxima_linha_vazia = row
                        break

            elif tipo_movimento == 'SAIDA':
                coluna_descricao = 8
                coluna_valor = 9
                coluna_forma_pg = 10
                coluna_tipo = 11

                for row in range(9, self.ws.max_row + 2):
                    linha_vazia = True
                    for col in range(8, 11):  # Colunas B (2) a G (7)
                        cell = self.ws.cell(row=row, column=col)
                        if cell.value is not None:
                            linha_vazia = False
                            break
                    if linha_vazia:
                        proxima_linha_vazia = row
                        break

            data_f = str(data).split('-')
            self.criar_planilha_caixa(data_f[0], data_f[1][1])

            if "tipo" in item:
                self.ws.cell(row=proxima_linha_vazia, column=coluna_tipo, value=item['tipo'])

            if "forma_pg" in item:
                self.ws.cell(row=proxima_linha_vazia, column=coluna_forma_pg, value=item['forma_pg'])

            if "descricao" in item:
                self.ws.cell(row=proxima_linha_vazia, column=coluna_descricao, value=descricao)

            if "valor" in item:
                self.ws.cell(row=proxima_linha_vazia, column=coluna_valor, value=item['valor'])

            for chave, valor in dicionario.items():
                if re.match(r'id\d+', chave):
                    lista_ids_audit_log.append(valor)

            self.wb_caixa.save(r"C:\Users\Igorj\OneDrive\Área de Trabalho\Planilhas\Caixa\2024\MAIO.xlsx")

            return lista_ids_audit_log

    def selecionar_sheet_por_data(self, data, tabela):
        """
        Seleciona a sheet com base na data fornecida.

        :param data: Data no formato 'dd-mm-yyyy'.
        """

        if tabela == 'caixa':
            data = str(data).split('-')
            nome_sheet = f'{data[2]}.{data[1][1]}'
            if nome_sheet in self.wb_caixa.sheetnames:
                self.ws = self.wb_caixa[nome_sheet]
                st.write(nome_sheet)
            else:
                print(f"Sheet '{nome_sheet}' não encontrada.")
                self.ws = None
        else:
            data = str(data).split('-')
            nome_sheet = f'{data[2]}.{data[1][1]}'
            if nome_sheet in self.wb.sheetnames:
                self.ws = self.wb[nome_sheet]
            else:
                print(f"Sheet '{nome_sheet}' não encontrada.")
                self.ws = None

    def insert_excel(self, dados_cliente):
        lista_ids_audit_log = []
        for nome, item in dados_cliente.items():
            st.write(nome)
            st.write(item)
            data = item['data']
            self.selecionar_sheet_por_data(data, 'reserva')

            if self.ws is None:
                print('Erro - Sheet nao selecionada')
                return

            coluna_nome_cliente = 2
            coluna_telefone = 3
            coluna_comissario = 4
            coluna_cert = 5
            coluna_roupa = 8
            coluna_receber_loja = 9
            coluna_pix = 10
            coluna_debito = 12
            coluna_credito = 14
            coluna_dinheiro = 16
            coluna_voucher = 18

            proxima_linha_vazia = None

            for row in range(6, self.ws.max_row + 2):
                linha_vazia = True
                for col in range(2, 8):  # Colunas B (2) a G (7)
                    cell = self.ws.cell(row=row, column=col)
                    if cell.value is not None:
                        linha_vazia = False
                        break
                if linha_vazia:
                    proxima_linha_vazia = row
                    break

            if proxima_linha_vazia is not None:
                if 'nome_cliente' in item:
                    self.ws.cell(row=proxima_linha_vazia, column=coluna_nome_cliente, value=item['nome_cliente'])
                if 'telefone' in item:
                    self.ws.cell(row=proxima_linha_vazia, column=coluna_telefone, value=item['telefone'])
                if 'nome_vendedor' in item:
                    self.ws.cell(row=proxima_linha_vazia, column=coluna_comissario,
                                 value=str(item['nome_vendedor']).upper())
                if 'tipo' in item:
                    self.ws.cell(row=proxima_linha_vazia, column=coluna_cert, value=item['tipo'])
                if 'roupa' in item:
                    self.ws.cell(row=proxima_linha_vazia, column=coluna_roupa, value=item['roupa'])
                if 'receber_loja':
                    self.ws.cell(row=proxima_linha_vazia, column=coluna_receber_loja, value=item['receber_loja'])

                if 'forma_pg' in item:
                    if item['recebedor'] == 'AcquaWorld':
                        if item['forma_pg'] == 'Pix':
                            self.ws.cell(row=proxima_linha_vazia, column=coluna_pix, value=item['pagamento'])

                        if item['forma_pg'] == 'Debito':
                            self.ws.cell(row=proxima_linha_vazia, column=coluna_debito, value=item['pagamento'])

                        if item['forma_pg'] == 'Credito':
                            self.ws.cell(row=proxima_linha_vazia, column=coluna_credito, value=item['pagamento'])

                        if item['forma_pg'] == 'Dinheiro':
                            self.ws.cell(row=proxima_linha_vazia, column=coluna_dinheiro, value=item['pagamento'])
                    else:
                        self.ws.cell(row=proxima_linha_vazia, column=coluna_voucher, value=item['pagamento'])

                # Lista para armazenar os valores das chaves que começam com "id" seguido por um número
                for chave, valor in item.items():
                    if re.match(r'id\d+', chave):
                        lista_ids_audit_log.append(valor)

                proxima_linha_vazia += 1
            else:
                print('Não foi encontrado linha vazia na coluna B')

        self.wb.save(r"C:\Users\Igorj\Downloads\MAIO 2024-teste.xlsx")

        return lista_ids_audit_log

    def encontrar_linha_pelo_nome(self, nome_cliente, coluna):
        if self.ws is None:
            print("Erro: Sheet não selecionada.")
            return None

        for row in range(6, self.ws.max_row + 1):  # Começa a partir da linha 6
            cell = self.ws.cell(row=row, column=coluna)
            if cell.value == nome_cliente:
                return row

        print(f"Nome '{nome_cliente}' não encontrado na coluna B.")
        return None

    def alterar_informacao(self, nome, data, coluna, novo_valor):

        num_colunas = {
            'coluna_nome_cliente': 2,
            'coluna_telefone': 3,
            'coluna_comissario': 4,
            'coluna_cert': 5,
            'coluna_roupa': 8
        }

        coluna = num_colunas[f'coluna_{coluna}']

        self.selecionar_sheet_por_data(data, 'reserva')

        if self.ws is None:
            print('Erro - Sheet nao selecionada')
            return

        row = self.encontrar_linha_pelo_nome(nome, 2)

        self.ws.cell(row=row, column=coluna, value=novo_valor)
        self.wb.save(self.caminho)

        print('Dados Alterados com Sucesso')

    def deletar_linha(self, nome, data):

        self.selecionar_sheet_por_data(data, reserva)
        row = self.encontrar_linha_pelo_nome(nome, 2)
        row_final = self.encontrar_linha_pelo_nome('STAFFS', 3)

        self.ws.delete_rows(row)
        row_final = row_final - 1
        self.ajustar_contagem(row_final)
        self.wb.save(self.caminho)
        print('Linha Excluida com Sucesso!')

    def ajustar_contagem(self, ultima_linha):

        contagem = 1
        for row in range(6, ultima_linha):
            self.ws.cell(row=row, column=1, value=contagem)  # Atualizar a contagem na coluna A
            contagem += 1

    def criar_planilha(self, ano, mes):

        sheet_modelo = self.wb_modelo['Modelo']
        num_dias = calendar.monthrange(ano, mes)[1]

        for dia in range(1, num_dias + 1):
            nome_sheet = f'{dia}.{mes}'

            # Copiar a sheet modelo
            nova_sheet = self.wb_modelo.copy_worksheet(sheet_modelo)
            nova_sheet.title = nome_sheet

        self.wb_modelo.save(r"C:\Users\Igorj\Downloads\MAIO 2024-teste.xlsx")

        print(num_dias)

    def criar_planilha_caixa(self, ano, mes):

        sheet_modelo = self.wb_modelo_caixa['Modelo']
        st.write(ano)
        st.write(mes)
        num_dias = calendar.monthrange(int(ano), int(mes))[1]
        output_path = r"C:\Users\Igorj\OneDrive\Área de Trabalho\Planilhas\Caixa\2024\MAIO.xlsx"

        # Verifica se a planilha já existe
        if os.path.exists(output_path):
            wb_saida = load_workbook(output_path)
        else:
            wb_saida = self.wb_modelo_caixa

        for dia in range(1, num_dias + 1):
            nome_sheet = f'{dia}.{mes}'

            # Verificar se a sheet já existe no workbook de saída
            if nome_sheet not in wb_saida.sheetnames:
                # Copiar a sheet modelo
                nova_sheet = wb_saida.copy_worksheet(sheet_modelo)
                nova_sheet.title = nome_sheet

        # Salvar o workbook atualizado
        wb_saida.save(output_path)
